"""
Size Breakdown Check -> Excel converter
=======================================
A Streamlit app that converts Haddad Apparel Group "SIZE BREAKDOWN CHECK"
(program EDIPAKLE) PDF reports into a clean Excel workbook.

The report is a fixed-width monospace layout, so every field sits at a stable
horizontal position. We parse by x-coordinate (via pdfplumber) rather than by
splitting on whitespace, which keeps things robust against blank cells and
variable-length size lists.

SIZE RATIOS layout (per the requested format):
  - Each record occupies two rows.
  - Row 1 holds all the fields plus  "ORDER: <values>"      in SIZE RATIOS.
  - Row 2 (fields blank) holds        "PRODUCED: <values>"  in SIZE RATIOS.
  - Multiple values are joined with dashes, e.g. 20-39-38-19.
  - Records are separated by an optional blank row.

Run with:  streamlit run app.py
Requires:  streamlit, pdfplumber, openpyxl, pandas
"""

import io
import re
import pdfplumber
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Report geometry (template constants for this specific report type).
# x-coordinates are in PDF points. Numbers are right-aligned inside the size
# slots, so we key off each token's right edge (x1).
# ---------------------------------------------------------------------------
FIELD_BOUNDS = [
    ("ACC #",       30, 58),
    ("STORE",       58, 84),
    ("PO#",         84, 148),
    ("RECEIVED",   148, 184),
    ("IBM",        184, 218),   # IBM number    ─┐ combined into "IBM/PCK"
    ("PCK",        218, 233),   # -N pick suffix ┘
    ("LIN#",       233, 247),
    ("SEASON/YR",  247, 272),
    ("DIV",        272, 281),
    ("STYLE",      281, 310),
    ("COLOR",      310, 326),
    ("LABEL",      326, 340),
]

SIZE_X0_MIN = 365      # left edge of the size-ratio region (after the Order/Produced label)
SIZE_X1_MAX = 665      # right edge (excludes the static "SizeRatio_Difference" label)
FIRST_SLOT_X1 = 391.8  # right edge of size slot #1
SLOT_PITCH = 25.2      # horizontal distance between adjacent size slots

BASE_COLUMNS = ["ACC #", "STORE", "PO#", "RECEIVED", "IBM/PCK", "LIN#",
                "SEASON/YR", "DIV", "STYLE", "COLOR", "LABEL"]
SIZE_COLUMN = "SIZE RATIOS"
ALL_COLUMNS = BASE_COLUMNS + [SIZE_COLUMN]

# Columns kept as text so Excel doesn't strip leading zeros / reinterpret values.
TEXT_COLUMNS = {"ACC #", "STORE", "PO#", "RECEIVED", "IBM/PCK", "SEASON/YR",
                "DIV", "STYLE", "COLOR", "LABEL", "SIZE RATIOS"}


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------
def _to_int(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return v


def _fmt_date(s):
    """Expand a 2-digit year to 4 digits: 6/24/26 -> 6/24/2026."""
    s = (s or "").strip()
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2})$", s)
    if m:
        mo, d, y = m.groups()
        return f"{int(mo)}/{int(d)}/20{y}"
    return s


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------
def _cluster_lines(words, tol=3):
    """Group words that share (roughly) the same vertical position into lines."""
    lines = []
    for w in sorted(words, key=lambda x: x["top"]):
        if lines and abs(w["top"] - lines[-1][0]) <= tol:
            lines[-1][1].append(w)
        else:
            lines.append([w["top"], [w]])
    return [sorted(ws, key=lambda x: x["x0"]) for _, ws in lines]


def _field_of(x0):
    for name, lo, hi in FIELD_BOUNDS:
        if lo <= x0 < hi:
            return name
    return None


def _make_slot_mapper(all_x1):
    """
    Map a token's right edge to a 0-based size slot, used only to order the
    numbers left-to-right. Uses the fixed template grid; if a page's
    coordinates don't line up (different margin/scale), anchors on the
    left-most observed value instead and flags it.
    """
    def fixed(x1):
        return round((x1 - FIRST_SLOT_X1) / SLOT_PITCH)

    drift_ok = True
    if all_x1:
        resid = max(abs(x1 - (FIRST_SLOT_X1 + SLOT_PITCH * fixed(x1))) for x1 in all_x1)
        drift_ok = resid <= 3.0

    if drift_ok:
        return fixed, True

    anchor = min(all_x1) if all_x1 else FIRST_SLOT_X1
    return (lambda x1: round((x1 - anchor) / SLOT_PITCH)), False


def parse_pdf(file_obj):
    """Return (records, grid_aligned). Each record is a dict of base fields plus
    '_order' / '_prod' mappings {slot_index: value_string}."""
    records = []
    all_x1 = []

    with pdfplumber.open(file_obj) as pdf:
        pages_lines = []
        for page in pdf.pages:
            lines = _cluster_lines(page.extract_words())
            pages_lines.append(lines)
            for ln in lines:
                texts = [w["text"] for w in ln]
                # Only sample size positions from real data lines (Order/Produced),
                # never from the header's "----- SIZE RATIOS -----" banner.
                if "Order" not in texts and "Produced" not in texts:
                    continue
                for w in ln:
                    if (w["text"] not in ("Order", "Produced")
                            and SIZE_X0_MIN <= w["x0"] and w["x1"] < SIZE_X1_MAX):
                        all_x1.append(w["x1"])

    slot_for, grid_aligned = _make_slot_mapper(all_x1)

    for lines in pages_lines:
        for i, ln in enumerate(lines):
            texts = [w["text"] for w in ln]
            if "Order" not in texts:          # only "Order" lines start a record
                continue

            rec = {n: "" for n, _, _ in FIELD_BOUNDS}
            order_slots = {}
            for w in ln:
                if w["text"] == "Order":
                    continue
                if SIZE_X0_MIN <= w["x0"] and w["x1"] < SIZE_X1_MAX:
                    order_slots[slot_for(w["x1"])] = w["text"]
                else:
                    f = _field_of(w["x0"])
                    if f:
                        rec[f] = (rec[f] + " " + w["text"]).strip()

            # The matching "Produced" line is the next line below (before any
            # following "Order" line) that contains the word Produced.
            prod_slots = {}
            for ln2 in lines[i + 1:]:
                t2 = [w["text"] for w in ln2]
                if "Order" in t2:
                    break
                if "Produced" in t2:
                    for w in ln2:
                        if (w["text"] != "Produced"
                                and SIZE_X0_MIN <= w["x0"] and w["x1"] < SIZE_X1_MAX):
                            prod_slots[slot_for(w["x1"])] = w["text"]
                    break

            ibm = rec.pop("IBM")
            pck = rec.pop("PCK")
            rec["IBM/PCK"] = " // ".join(p for p in (ibm, pck) if p)
            rec["RECEIVED"] = _fmt_date(rec["RECEIVED"])
            rec["_order"] = order_slots
            rec["_prod"] = prod_slots
            records.append(rec)

    return records, grid_aligned


# ---------------------------------------------------------------------------
# Shaping records into the two-row ORDER / PRODUCED layout
# ---------------------------------------------------------------------------
def _join_ratios(slots):
    return "-".join(slots[k] for k in sorted(slots))


def build_rows(records, separator=True):
    """
    Expand records into output rows. Each record -> an ORDER row (with all the
    fields) and a PRODUCED row (fields blank), optionally followed by a blank
    separator row. The 'kind' key marks each row for styling and is not output.
    """
    rows = []
    for idx, r in enumerate(records):
        order_row = {c: r[c] for c in BASE_COLUMNS}
        order_row["LIN#"] = _to_int(r["LIN#"])
        order_row[SIZE_COLUMN] = f"ORDER: {_join_ratios(r['_order'])}"
        order_row["kind"] = "order"
        rows.append(order_row)

        prod_row = {c: "" for c in BASE_COLUMNS}
        prod_row[SIZE_COLUMN] = f"PRODUCED: {_join_ratios(r['_prod'])}"
        prod_row["kind"] = "produced"
        rows.append(prod_row)

        if separator and idx < len(records) - 1:
            blank = {c: "" for c in ALL_COLUMNS}
            blank["kind"] = "blank"
            rows.append(blank)
    return rows


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build_workbook(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Size Breakdown"

    # Header row
    for c, header in enumerate(ALL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = BORDER

    # Data rows
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, header in enumerate(ALL_COLUMNS, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=row.get(header, ""))
            cell.border = BORDER
            if header in TEXT_COLUMNS:
                cell.number_format = "@"
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    # Filter dropdown buttons across the whole table.
    last_row = len(rows) + 1
    last_col = get_column_letter(len(ALL_COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # Auto-fit each column to the widest of its header (plus room for the
    # filter-dropdown button) and its cell contents.
    for c_idx, header in enumerate(ALL_COLUMNS, start=1):
        col = get_column_letter(c_idx)
        content_len = max(
            [len(str(row.get(header, ""))) for row in rows
             if row.get(header, "") not in (None, "")] + [0]
        )
        width = max(content_len, len(str(header)) + 3) + 1
        ws.column_dimensions[col].width = min(width, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------------------------
st.set_page_config(page_title="Size Breakdown -> Excel", page_icon="📐", layout="wide")
st.title("Size Breakdown Check → Excel")
st.caption("Converts Haddad Apparel Group SIZE BREAKDOWN CHECK (EDIPAKLE) PDFs into a clean spreadsheet.")

uploaded = st.file_uploader(
    "Upload one or more Size Breakdown PDFs",
    type=["pdf"],
    accept_multiple_files=True,
)

separator = st.checkbox("Add a blank row between records", value=True)

if uploaded:
    all_records = []
    grid_warned = False
    for f in uploaded:
        try:
            recs, aligned = parse_pdf(f)
        except Exception as e:  # noqa: BLE001
            st.error(f"Could not read **{f.name}**: {e}")
            continue
        if not recs:
            st.warning(f"No Size Breakdown records found in **{f.name}** — is it the right report type?")
        if not aligned and not grid_warned:
            st.info("This PDF's size columns didn't match the expected grid exactly, so the "
                    "left-to-right order was inferred from the data. Spot-check the SIZE RATIOS.")
            grid_warned = True
        all_records.extend(recs)

    if all_records:
        rows = build_rows(all_records, separator=separator)
        df = pd.DataFrame(rows, columns=ALL_COLUMNS).fillna("")
        st.success(f"Parsed {len(all_records)} record(s) from {len(uploaded)} file(s).")
        st.dataframe(df, use_container_width=True, hide_index=True)

        st.download_button(
            "⬇️ Download Excel",
            data=build_workbook(rows),
            file_name="size_breakdown.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
else:
    st.info("Upload a PDF to get started.")
